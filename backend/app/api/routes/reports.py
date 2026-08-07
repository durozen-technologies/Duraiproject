import io
from datetime import date
from pathlib import Path
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ...db.database import get_db
from ...models.purchase import Purchase
from ...models.sale import Sale
from ...models.party import Party
from ...models.transaction import PaymentTransaction
from ...models.enums import TransactionType

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

router = APIRouter(prefix="/reports", tags=["reports"])

_UNICODE_FONT: Optional[str] = None
_UNICODE_FONT_TRIED = False


def _fmt_inr(amount: float) -> str:
    """Indian grouping with 2 decimals, e.g. 1,02,000.00"""
    n = float(amount or 0)
    sign = "-" if n < 0 else ""
    n = abs(n)
    whole, frac = f"{n:.2f}".split(".")
    if len(whole) <= 3:
        grouped = whole
    else:
        last3 = whole[-3:]
        rest = whole[:-3]
        parts = []
        while rest:
            parts.append(rest[-2:])
            rest = rest[:-2]
        grouped = ",".join(reversed(parts)) + "," + last3
    return f"{sign}{grouped}.{frac}"


def _fmt_weight(value: float) -> str:
    return f"{float(value or 0):.2f}"


def _fmt_rate(value: float) -> str:
    return f"Rs.{_fmt_inr(value)}"


def _display_bill_number(bill_number: Optional[str]) -> str:
    """Show YYYY-00000x without DPS- prefix for report cells."""
    raw = (bill_number or "").strip()
    if not raw:
        return "-"
    upper = raw.upper()
    if upper.startswith("DPS-"):
        raw = raw[4:]
    return raw or "-"


def _bill_no_cell(bill_number: Optional[str], font_size: int = 7):
    """Paragraph so bill no wraps inside the column instead of overflowing."""
    text = _display_bill_number(bill_number)
    if text == "-":
        return "-"
    style = ParagraphStyle(
        "BillNoCell",
        fontName="Helvetica",
        fontSize=font_size,
        leading=font_size + 1,
        alignment=1,  # CENTER
    )
    safe = text.replace("&", "&amp;").replace("<", "&lt;")
    return Paragraph(safe, style)


def _normalize_lang(language: Optional[str]) -> str:
    lang = (language or "en").strip().lower()
    return "ta" if lang in ("ta", "tamil") else "en"


def _party_display_name(party: Optional[Party], language: str) -> str:
    if not party:
        return "Unknown"
    if language == "ta":
        tamil = (getattr(party, "tamil_name", None) or "").strip()
        if tamil:
            return tamil
    return (party.name or "").strip() or "Unknown"


def _has_tamil_text(value: str) -> bool:
    return any("\u0b80" <= ch <= "\u0bff" for ch in (value or ""))


def _unicode_font_name() -> Optional[str]:
    """Register Noto Sans Tamil (Duro_POS-style) once for PDF Tamil text."""
    global _UNICODE_FONT, _UNICODE_FONT_TRIED
    if _UNICODE_FONT_TRIED:
        return _UNICODE_FONT
    _UNICODE_FONT_TRIED = True

    # reports.py -> routes -> api -> app
    app_dir = Path(__file__).resolve().parents[2]
    bundled = app_dir / "assets" / "fonts"
    candidates: list[tuple[Path, Optional[int]]] = [
        (bundled / "NotoSansTamil-Regular.ttf", None),
        (bundled / "NotoSansTamil.ttf", None),
        (bundled / "Nirmala.ttf", None),
        (Path(r"C:\Windows\Fonts\Nirmala.ttf"), None),
        (Path(r"C:\Windows\Fonts\Nirmala.ttc"), 0),
        (Path(r"C:\Windows\Fonts\Nirmala.ttc"), 1),
        (Path(r"C:\Windows\Fonts\latha.ttf"), None),
        (Path("/usr/share/fonts/truetype/noto/NotoSansTamil-Regular.ttf"), None),
        (Path("/usr/share/fonts/truetype/noto/NotoSansTamilUI-Regular.ttf"), None),
        (Path("/usr/share/fonts/truetype/noto/NotoSerifTamil-Regular.ttf"), None),
    ]
    for path, subfont_index in candidates:
        if not path.exists():
            continue
        try:
            kwargs = {}
            if subfont_index is not None:
                kwargs["subfontIndex"] = subfont_index
            pdfmetrics.registerFont(TTFont("ReportUnicode", str(path), **kwargs))
            # Smoke-check Tamil glyph metrics (avoids silent broken faces)
            if pdfmetrics.stringWidth("அ", "ReportUnicode", 10) <= 0:
                continue
            _UNICODE_FONT = "ReportUnicode"
            return _UNICODE_FONT
        except Exception:
            continue
    return None


def _party_name_cell(name: str, language: str, font_size: int = 8):
    """Paragraph with Tamil-capable font whenever Tamil text is present."""
    use_tamil = language == "ta" or _has_tamil_text(name)
    font = _unicode_font_name() if use_tamil else None
    if not font:
        return name
    style = ParagraphStyle(
        "PartyNameCell",
        fontName=font,
        fontSize=font_size,
        leading=font_size + 3,
    )
    safe = (name or "").replace("&", "&amp;").replace("<", "&lt;")
    return Paragraph(safe, style)


def _party_nickname(party: Optional[Party]) -> str:
    if not party:
        return ""
    return (getattr(party, "nickname", None) or "").strip()


def _party_excel_cell(party: Optional[Party], language: str) -> str:
    """Party/purchaser name with nickname on the next line (same Excel cell)."""
    name = _party_display_name(party, language)
    nick = _party_nickname(party)
    if nick:
        return f"{name}\n{nick}"
    return name


def _party_pdf_cell(party: Optional[Party], language: str, font_size: int = 8):
    """PDF cell: bold party name with nickname underneath."""
    name = _party_display_name(party, language)
    nick = _party_nickname(party)
    use_tamil = language == "ta" or _has_tamil_text(name) or _has_tamil_text(nick)
    font = _unicode_font_name() if use_tamil else "Helvetica"
    if not font:
        font = "Helvetica"
    safe_name = (name or "Unknown").replace("&", "&amp;").replace("<", "&lt;")
    if nick:
        safe_nick = nick.replace("&", "&amp;").replace("<", "&lt;")
        nick_size = max(font_size - 1, 6)
        html = (
            f"<b>{safe_name}</b><br/>"
            f"<font size='{nick_size}' color='#555555'>{safe_nick}</font>"
        )
    else:
        html = f"<b>{safe_name}</b>"
    style = ParagraphStyle(
        "PartyNameNickCell",
        fontName=font,
        fontSize=font_size,
        leading=font_size + 2,
    )
    return Paragraph(html, style)


def _driver_name_from_row(row) -> str:
    driver = getattr(row, "driver", None)
    if driver and (getattr(driver, "name", None) or "").strip():
        return driver.name.strip()
    return (getattr(row, "driver_name", None) or "").strip() or "-"


def _driver_mobile_from_row(row) -> str:
    driver = getattr(row, "driver", None)
    if not driver:
        return ""
    return (getattr(driver, "mobile", None) or "").strip()


def _driver_excel_cell(row) -> str:
    """Driver name with mobile on the next line (same Excel cell)."""
    name = _driver_name_from_row(row)
    mobile = _driver_mobile_from_row(row)
    if name == "-":
        return "-"
    if mobile:
        return f"{name}\n{mobile}"
    return name


def _driver_pdf_cell(row, font_size: int = 7):
    """PDF cell: bold driver name with mobile underneath."""
    name = _driver_name_from_row(row)
    mobile = _driver_mobile_from_row(row)
    if name == "-" and not mobile:
        return "-"
    safe_name = (name or "-").replace("&", "&amp;").replace("<", "&lt;")
    if mobile:
        safe_mobile = mobile.replace("&", "&amp;").replace("<", "&lt;")
        mobile_size = max(font_size - 1, 6)
        html = (
            f"<b>{safe_name}</b><br/>"
            f"<font size='{mobile_size}' color='#555555'>{safe_mobile}</font>"
        )
    else:
        html = f"<b>{safe_name}</b>"
    style = ParagraphStyle(
        "DriverNameCell",
        fontName="Helvetica",
        fontSize=font_size,
        leading=font_size + 2,
    )
    return Paragraph(html, style)


def _party_title_paragraph(name: str, language: str, base_style: ParagraphStyle) -> Paragraph:
    use_tamil = language == "ta" or _has_tamil_text(name)
    font = _unicode_font_name() if use_tamil else None
    style = ParagraphStyle(
        "PartyTitleUnicode",
        parent=base_style,
        fontName=font or base_style.fontName,
        fontSize=base_style.fontSize,
        leading=(base_style.fontSize or 12) + 4,
    )
    safe = (name or "").replace("&", "&amp;").replace("<", "&lt;")
    return Paragraph(safe if use_tamil else safe.upper(), style)


def _table_body_font(language: str) -> str:
    """Latin body font for numeric/ASCII cells. Tamil names use Paragraph + unicode font."""
    return "Helvetica"


def _normalize_export_format(fmt: Optional[str]) -> str:
    value = (fmt or "pdf").strip().lower()
    if value in ("xlsx", "excel", "xls"):
        return "xlsx"
    return "pdf"


def _excel_response(
    *,
    title: str,
    headers: list[str],
    rows: list[list],
    filename: str,
    sheet_name: str = "Report",
    summary_rows: Optional[list[list]] = None,
) -> StreamingResponse:
    """Build an .xlsx workbook and return it as a download response."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Report")[:31]

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10)
    thin = Border(
        left=Side(style="thin", color="CFCFCF"),
        right=Side(style="thin", color="CFCFCF"),
        top=Side(style="thin", color="CFCFCF"),
        bottom=Side(style="thin", color="CFCFCF"),
    )

    col_count = max(len(headers), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    for r_idx, row in enumerate(rows, header_row + 1):
        for c_idx in range(1, col_count + 1):
            value = row[c_idx - 1] if c_idx - 1 < len(row) else ""
            cell = ws.cell(r_idx, c_idx, value)
            cell.border = thin
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.00"
            else:
                wrap = isinstance(value, str) and "\n" in value
                cell.alignment = Alignment(
                    horizontal="left" if c_idx in (2, 5) else "center",
                    vertical="center",
                    wrap_text=wrap,
                )

    next_row = header_row + 1 + len(rows) + 1
    if summary_rows:
        summary_fill = PatternFill("solid", fgColor="E8F5EE")
        for offset, summary in enumerate(summary_rows):
            for c_idx in range(1, col_count + 1):
                value = summary[c_idx - 1] if c_idx - 1 < len(summary) else ""
                cell = ws.cell(next_row + offset, c_idx, value)
                cell.border = thin
                cell.font = Font(bold=True)
                cell.fill = summary_fill
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

    for col in range(1, col_count + 1):
        max_len = 12
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value is None:
                    continue
                max_len = max(max_len, min(42, len(str(cell.value))))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    ws.row_dimensions[header_row].height = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
@router.get("/purchases")
async def generate_purchase_report(
    from_date: date,
    to_date: date,
    party_id: Optional[UUID] = None,
    language: Optional[str] = Query("en", description="Party name language: en or ta"),
    format: Optional[str] = Query("pdf", description="Export format: pdf or xlsx"),
    db: AsyncSession = Depends(get_db)
):
    lang = _normalize_lang(language)
    export_fmt = _normalize_export_format(format)
    query = select(Purchase).options(
        selectinload(Purchase.party),
        selectinload(Purchase.driver),
    ).where(
        and_(Purchase.date >= from_date, Purchase.date <= to_date)
    ).order_by(Purchase.date)
    
    if party_id:
        query = query.where(Purchase.party_id == party_id)
        
    result = await db.execute(query)
    purchases = result.scalars().all()

    date_label = f"{from_date.strftime('%d-%m-%Y')} to {to_date.strftime('%d-%m-%Y')}"
    file_stem = f"purchase_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}"

    if export_fmt == "xlsx":
        headers = [
            "Date", "Purchaser Name", "Vehicle No", "Bill No", "Driver Name",
            "Total Box", "Total Birds", "Weighbridge Weight(Kg)", "Net Weight",
            "Avg Wt", "Rate (Rs.)", "Amount (Rs.)", "Paid Cash (Rs.)", "Paid UPI (Rs.)",
            "Total Paid Amount (Rs.)", "Balance (Rs.)",
        ]
        excel_rows = []
        for p in purchases:
            net_wt = float(p.net_weight or 0)
            birds = int(p.actual_birds or 0)
            calc_avg_wt = (net_wt / birds) if birds > 0 else 0.0
            cash = float(p.cash_payment or 0)
            upi = float(p.upi_payment or 0)
            excel_rows.append([
                p.date.strftime("%d-%m-%Y") if p.date else "-",
                _party_excel_cell(p.party, lang),
                p.vehicle_number or "-",
                _display_bill_number(p.bill_number),
                _driver_excel_cell(p),
                int(p.total_boxes or 0),
                birds,
                float(p.weighbridge_weight or 0),
                net_wt,
                round(calc_avg_wt, 2),
                float(p.purchase_rate or 0),
                float(p.purchase_amount or 0),
                cash,
                upi,
                round(cash + upi, 2),
                float(p.balance_amount or 0),
            ])
        return _excel_response(
            title=f"Purchase Register ({date_label})",
            headers=headers,
            rows=excel_rows,
            filename=f"{file_stem}.xlsx",
            sheet_name="Purchases",
        )
    
    # Generate PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15,
        leftMargin=15,
        topMargin=20,
        bottomMargin=20
    )
    
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    elements.append(Paragraph(f"Purchase Register ({date_label})", title_style))
    elements.append(Spacer(1, 10))
    
    # Table Header based on Purchase.html
    data = [
        [
            "Date", "Purchaser\nName", "Vehicle\nNo", "Bill\nNo", "Driver\nName",
            "Total\nBox", "Total\nBirds", "Weighbridge\nWeight(Kg)", "Net\nWeight",
            "Avg\nWt", "Rate\n(Rs.)", "Amount\n(Rs.)", "Paid Cash\n(Rs.)", "Paid UPI\n(Rs.)", "Total Paid\nAmount (Rs.)", "Balance\n(Rs.)"
        ]
    ]
    
    for p in purchases:
        # Convert rate and amounts to float, then format
        amount = f"{p.purchase_amount:,.2f}" if p.purchase_amount else "0.00"
        cash = f"{p.cash_payment:,.2f}" if p.cash_payment else "0.00"
        upi = f"{p.upi_payment:,.2f}" if p.upi_payment else "0.00"
        balance = f"{p.balance_amount:,.2f}" if p.balance_amount else "0.00"
        rate = f"{p.purchase_rate:,.2f}" if p.purchase_rate else "0.00"
        
        # Calculate Average Weight dynamically for the report
        net_wt = p.net_weight or 0
        birds = p.actual_birds or 0
        calc_avg_wt = (net_wt / birds) if birds > 0 else 0.0
        
        # Calculate Total Paid dynamically
        total_paid_val = (p.cash_payment or 0) + (p.upi_payment or 0)
        total_paid = f"{total_paid_val:,.2f}"
        
        row = [
            p.date.strftime('%d-%m-%Y') if p.date else "-",
            _party_pdf_cell(p.party, lang, font_size=8),
            p.vehicle_number or "-",
            _bill_no_cell(p.bill_number, font_size=8),
            _driver_pdf_cell(p, font_size=8),
            str(p.total_boxes) if p.total_boxes else "0",
            str(p.actual_birds) if p.actual_birds else "0",
            f"{p.weighbridge_weight:.2f}" if p.weighbridge_weight else "0.00",
            f"{p.net_weight:.2f}" if p.net_weight else "0.00",
            f"{calc_avg_wt:.2f}",
            rate,
            amount,
            cash,
            upi,
            total_paid,
            balance
        ]
        data.append(row)
        
    # Table Style
    t = Table(data, repeatRows=1)
    
    # Column widths based on A4 Landscape (842pts) minus 30pts margins = 812pts usable
    usable_width = 812
    t._argW = [
        usable_width * 0.07, # Date
        usable_width * 0.08, # Purchaser Name
        usable_width * 0.08, # Vehicle No
        usable_width * 0.08, # Bill No
        usable_width * 0.07, # Driver Name
        usable_width * 0.04, # Total Box
        usable_width * 0.05, # Total Birds
        usable_width * 0.07, # Weighbridge Weight
        usable_width * 0.05, # Net Weight
        usable_width * 0.04, # Avg Wt
        usable_width * 0.06, # Rate
        usable_width * 0.06, # Amount
        usable_width * 0.05, # Paid Cash
        usable_width * 0.05, # Paid UPI
        usable_width * 0.07, # Total Paid Amount
        usable_width * 0.08  # Balance
    ]
    
    body_font = _table_body_font(lang)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d9d9d9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cfcfcf')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#cfcfcf')),
        ('FONTNAME', (0, 1), (-1, -1), body_font),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Date
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Party Name (Paragraph keeps its own font)
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),  # Vehicle, Bill
        ('ALIGN', (4, 1), (4, -1), 'LEFT'),  # Driver Name
        ('ALIGN', (5, 1), (9, -1), 'CENTER'),  # Boxes to Avg Wt
        ('ALIGN', (10, 1), (-1, -1), 'RIGHT'),  # Rate to Balance
    ]
    if lang != "ta":
        style_cmds.append(('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'))
        style_cmds.append(('FONTNAME', (11, 1), (-1, -1), 'Helvetica-Bold'))
    
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={file_stem}.pdf"}
    )

@router.get("/sales")
async def generate_sale_report(
    from_date: date,
    to_date: date,
    party_id: Optional[UUID] = None,
    language: Optional[str] = Query("en", description="Party name language: en or ta"),
    format: Optional[str] = Query("pdf", description="Export format: pdf or xlsx"),
    db: AsyncSession = Depends(get_db)
):
    lang = _normalize_lang(language)
    export_fmt = _normalize_export_format(format)
    query = select(Sale).options(
        selectinload(Sale.party),
        selectinload(Sale.driver),
    ).where(
        and_(Sale.date >= from_date, Sale.date <= to_date)
    ).order_by(Sale.date)
    
    if party_id:
        query = query.where(Sale.party_id == party_id)
        
    result = await db.execute(query)
    sales = result.scalars().all()

    date_label = f"{from_date.strftime('%d-%m-%Y')} to {to_date.strftime('%d-%m-%Y')}"
    file_stem = f"sale_report_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}"

    if export_fmt == "xlsx":
        headers = [
            "Date", "Party Name", "Vehicle No", "Bill No", "Driver Name",
            "Weighbridge Weight(Kg)", "Net Weight(Kg)", "Weight Rate (Rs.)", "Net Wt Amount (Rs.)",
            "Box", "Box Rate (Rs.)", "Box Amount (Rs.)",
            "Total Amount (Rs.)", "Paid Cash (Rs.)", "Paid UPI (Rs.)", "Total Paid Amount (Rs.)", "Balance (Rs.)",
        ]
        excel_rows = []
        for s in sales:
            net_wt = float(s.weight or 0)
            wt_rate_val = float(s.weight_rate or 0)
            boxes = int(s.boxes or 0)
            box_rate_val = float(s.box_rate or 0)
            weight_amount = float(s.weight_amount or 0) or round(net_wt * wt_rate_val, 2)
            box_amount = float(s.box_amount or 0) or round(boxes * box_rate_val, 2)
            total_amount = float(s.total_invoice_amount or 0) or round(weight_amount + box_amount, 2)
            cash = float(s.cash_payment or 0)
            upi = float(s.upi_payment or 0)
            excel_rows.append([
                s.date.strftime("%d-%m-%Y") if s.date else "-",
                _party_excel_cell(s.party, lang),
                s.vehicle_number or "-",
                _display_bill_number(s.bill_number),
                _driver_excel_cell(s),
                float(s.weighbridge_weight or 0),
                net_wt,
                wt_rate_val,
                weight_amount,
                boxes,
                box_rate_val,
                box_amount,
                total_amount,
                cash,
                upi,
                round(cash + upi, 2),
                float(s.balance_amount or 0),
            ])
        return _excel_response(
            title=f"Sales Register ({date_label})",
            headers=headers,
            rows=excel_rows,
            filename=f"{file_stem}.xlsx",
            sheet_name="Sales",
        )
    
    # Generate PDF in memory
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15,
        leftMargin=15,
        topMargin=20,
        bottomMargin=20
    )
    
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = 1 # Center
    
    elements.append(Paragraph(f"Sales Register ({date_label})", title_style))
    elements.append(Spacer(1, 10))
    
    # Net kg × rate and Boxes × box rate as separate amount columns; total = sum
    data = [
        [
            "Date", "Party\nName", "Vehicle\nNo", "Bill\nNo", "Driver\nName",
            "Weighbridge\nWeight(Kg)", "Net\nWeight(Kg)", "Weight\nRate (Rs.)", "Net Wt\nAmount (Rs.)",
            "Box", "Box\nRate (Rs.)", "Box\nAmount (Rs.)",
            "Total\nAmount (Rs.)", "Paid Cash\n(Rs.)", "Paid UPI\n(Rs.)", "Total Paid\nAmount (Rs.)", "Balance\n(Rs.)"
        ]
    ]
    
    for s in sales:
        net_wt = float(s.weight or 0)
        wt_rate_val = float(s.weight_rate or 0)
        boxes = int(s.boxes or 0)
        box_rate_val = float(s.box_rate or 0)
        weight_amount = float(s.weight_amount or 0) or round(net_wt * wt_rate_val, 2)
        box_amount = float(s.box_amount or 0) or round(boxes * box_rate_val, 2)
        total_amount = float(s.total_invoice_amount or 0) or round(weight_amount + box_amount, 2)

        cash = f"{float(s.cash_payment or 0):,.2f}"
        upi = f"{float(s.upi_payment or 0):,.2f}"
        balance = f"{float(s.balance_amount or 0):,.2f}"
        total_paid_val = float(s.cash_payment or 0) + float(s.upi_payment or 0)
        total_paid = f"{total_paid_val:,.2f}"
        
        row = [
            s.date.strftime('%d-%m-%Y') if s.date else "-",
            _party_pdf_cell(s.party, lang, font_size=7),
            s.vehicle_number or "-",
            _bill_no_cell(s.bill_number, font_size=7),
            _driver_pdf_cell(s, font_size=7),
            f"{float(s.weighbridge_weight or 0):.2f}",
            f"{net_wt:.2f}",
            f"{wt_rate_val:,.2f}",
            f"{weight_amount:,.2f}",
            str(boxes),
            f"{box_rate_val:,.2f}",
            f"{box_amount:,.2f}",
            f"{total_amount:,.2f}",
            cash,
            upi,
            total_paid,
            balance,
        ]
        data.append(row)
        
    # Table Style
    t = Table(data, repeatRows=1)
    
    # A4 Landscape usable ~812pts; 17 columns
    usable_width = 812
    t._argW = [
        usable_width * 0.055,  # Date
        usable_width * 0.085,  # Party Name
        usable_width * 0.055,  # Vehicle No
        usable_width * 0.07,   # Bill No (YYYY-000001)
        usable_width * 0.06,   # Driver Name
        usable_width * 0.055,  # Weighbridge
        usable_width * 0.05,   # Net Weight
        usable_width * 0.05,   # Wt Rate
        usable_width * 0.055,  # Net Wt Amount
        usable_width * 0.035,  # Box
        usable_width * 0.05,   # Box Rate
        usable_width * 0.055,  # Box Amount
        usable_width * 0.06,   # Total Amt
        usable_width * 0.05,   # Paid Cash
        usable_width * 0.045,  # Paid UPI
        usable_width * 0.055,  # Total Paid
        usable_width * 0.055,  # Balance
    ]
    
    body_font = _table_body_font(lang)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d9d9d9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 1),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cfcfcf')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#cfcfcf')),
        ('FONTNAME', (0, 1), (-1, -1), body_font),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'LEFT'),
        ('ALIGN', (5, 1), (11, -1), 'CENTER'),
        ('ALIGN', (12, 1), (-1, -1), 'RIGHT'),
    ]
    if lang != "ta":
        style_cmds.append(('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'))
    
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={file_stem}.pdf"}
    )


def _balance_suffix(amount: float) -> str:
    return "Cr" if amount >= 0 else "Dr"


def _fmt_balance(amount: float) -> str:
    return f"Rs.{_fmt_inr(abs(amount))} {_balance_suffix(amount)}"


async def _party_payment_net(
    db: AsyncSession,
    party_id: UUID,
    *,
    before: Optional[date] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
) -> float:
    """Net payments that reduce credit balance: PAID − RECEIVED."""
    async def _sum_type(txn_type: TransactionType) -> float:
        filters = [
            PaymentTransaction.party_id == party_id,
            PaymentTransaction.type == txn_type,
        ]
        if before is not None:
            filters.append(PaymentTransaction.date < before)
        if from_date is not None:
            filters.append(PaymentTransaction.date >= from_date)
        if to_date is not None:
            filters.append(PaymentTransaction.date <= to_date)
        result = await db.execute(
            select(func.coalesce(func.sum(PaymentTransaction.total_amount), 0)).where(and_(*filters))
        )
        return float(result.scalar() or 0)

    paid = await _sum_type(TransactionType.PAID)
    received = await _sum_type(TransactionType.RECEIVED)
    return paid - received


@router.get("/party-ledger")
async def generate_party_ledger_report(
    party_id: UUID,
    from_date: date,
    to_date: date,
    language: Optional[str] = Query("en", description="Party name language: en or ta"),
    format: Optional[str] = Query("pdf", description="Export format: pdf or xlsx"),
    db: AsyncSession = Depends(get_db),
):
    """Party Ledger PDF/Excel matching Partystatement.html layout."""
    lang = _normalize_lang(language)
    export_fmt = _normalize_export_format(format)
    party_result = await db.execute(select(Party).where(Party.id == party_id))
    party = party_result.scalar_one_or_none()
    if not party:
        raise HTTPException(status_code=404, detail="Party not found")

    # Prior-period invoices + payments before from_date, plus party opening
    # e.g. opening 1000 + July invoices net − July paid = statement opening for Aug
    prior_credit_result = await db.execute(
        select(func.coalesce(func.sum(Purchase.purchase_amount), 0)).where(
            and_(Purchase.party_id == party_id, Purchase.date < from_date)
        )
    )
    prior_debit_result = await db.execute(
        select(func.coalesce(func.sum(Sale.total_invoice_amount), 0)).where(
            and_(Sale.party_id == party_id, Sale.date < from_date)
        )
    )
    prior_credit = float(prior_credit_result.scalar() or 0)
    prior_debit = float(prior_debit_result.scalar() or 0)
    prior_paid_net = await _party_payment_net(db, party_id, before=from_date)
    party_opening = float(party.opening_balance or 0)
    opening_balance = party_opening + (prior_credit - prior_debit) - prior_paid_net

    purchases_result = await db.execute(
        select(Purchase)
        .where(
            and_(
                Purchase.party_id == party_id,
                Purchase.date >= from_date,
                Purchase.date <= to_date,
            )
        )
        .order_by(Purchase.date.asc(), Purchase.created_at.asc())
    )
    purchases = purchases_result.scalars().all()

    sales_result = await db.execute(
        select(Sale)
        .where(
            and_(
                Sale.party_id == party_id,
                Sale.date >= from_date,
                Sale.date <= to_date,
            )
        )
        .order_by(Sale.date.asc(), Sale.created_at.asc())
    )
    sales = sales_result.scalars().all()

    rows: list[dict] = []
    for p in purchases:
        rows.append(
            {
                "date": p.date,
                "kind": "Purchase",
                "net_weight": float(p.net_weight or 0),
                "rate": float(p.purchase_rate or 0),
                "credit": float(p.purchase_amount or 0),
                "debit": 0.0,
                "sort_key": (p.date, 0, str(p.id)),
            }
        )
    for s in sales:
        rows.append(
            {
                "date": s.date,
                "kind": "Sale",
                "net_weight": float(s.weight or 0),
                "rate": float(s.weight_rate or 0),
                "credit": 0.0,
                "debit": float(s.total_invoice_amount or 0),
                "sort_key": (s.date, 1, str(s.id)),
            }
        )
    rows.sort(key=lambda r: r["sort_key"])

    total_credit = sum(r["credit"] for r in rows)
    total_debit = sum(r["debit"] for r in rows)
    difference = total_credit - total_debit
    # Period payments (PAID − RECEIVED); subtracted from opening + difference
    paid_amount = await _party_payment_net(db, party_id, from_date=from_date, to_date=to_date)
    # Closing must match the party's live current balance
    closing_balance = float(party.current_balance or 0)

    display_name = _party_display_name(party, lang)
    date_label = f"{from_date.strftime('%d-%m-%Y')} to {to_date.strftime('%d-%m-%Y')}"
    safe_name = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in (party.name or "party"))[:40]
    file_stem = f"party_ledger_{safe_name}_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}"
    opening_label = _fmt_balance(opening_balance)
    closing_label = _fmt_balance(closing_balance)
    paid_label = f"Rs.{_fmt_inr(paid_amount)}"

    if export_fmt == "xlsx":
        headers = ["Date", "Type", "Net Weight (Kg)", "Rate (Rs.)", "Credit (Rs.)", "Debit (Rs.)"]
        excel_rows = []
        for r in rows:
            is_purchase = r["kind"] == "Purchase"
            excel_rows.append([
                r["date"].strftime("%d-%m-%Y") if r["date"] else "-",
                r["kind"],
                round(float(r["net_weight"]), 2),
                round(float(r["rate"]), 2),
                round(float(r["credit"]), 2) if is_purchase else 0,
                round(float(r["debit"]), 2) if not is_purchase else 0,
            ])
        summary = [
            ["", "", "", "Opening Balance", opening_label, ""],
            ["", "", "", "Total Credit", round(total_credit, 2), ""],
            ["", "", "", "Total Debit", "", round(total_debit, 2)],
            ["", "", "", "Difference", round(difference, 2), ""],
            ["", "", "", "Paid Amount", round(paid_amount, 2), ""],
            ["", "", "", "Closing Balance", closing_label, ""],
        ]
        return _excel_response(
            title=f"{display_name} — Party Ledger ({date_label}) | Opening Balance: {opening_label}",
            headers=headers,
            rows=excel_rows,
            filename=f"{file_stem}.xlsx",
            sheet_name="Party Ledger",
            summary_rows=summary,
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PartyLedgerTitle",
        parent=styles["Heading1"],
        alignment=1,
        fontSize=22,
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "PartyLedgerSubtitle",
        parent=styles["Normal"],
        alignment=1,
        fontSize=12,
        textColor=colors.HexColor("#666666"),
        spaceAfter=4,
    )
    range_style = ParagraphStyle(
        "PartyLedgerRange",
        parent=styles["Normal"],
        alignment=1,
        fontSize=11,
        textColor=colors.HexColor("#444444"),
        spaceAfter=8,
    )
    opening_style = ParagraphStyle(
        "PartyLedgerOpening",
        parent=styles["Normal"],
        alignment=2,  # RIGHT — above table header, right corner
        fontSize=11,
        textColor=colors.HexColor("#111111"),
        spaceAfter=10,
    )

    elements = []
    elements.append(_party_title_paragraph(display_name, lang, title_style))
    elements.append(Paragraph("Party Ledger Statement", subtitle_style))
    elements.append(Paragraph(date_label, range_style))
    elements.append(
        Paragraph(f"<b>Opening Balance:</b> {opening_label}", opening_style)
    )

    table_data = [
        ["Date", "Type", "Net Weight (Kg)", "Rate", "Credit (Rs.)", "Debit (Rs.)"]
    ]

    green = colors.HexColor("#0b8b32")
    red = colors.HexColor("#d10000")

    for r in rows:
        is_purchase = r["kind"] == "Purchase"
        credit_cell = f"Rs.{_fmt_inr(r['credit'])}" if is_purchase else "-"
        debit_cell = f"Rs.{_fmt_inr(r['debit'])}" if not is_purchase else "-"
        table_data.append(
            [
                r["date"].strftime("%d-%m-%Y") if r["date"] else "-",
                r["kind"],
                _fmt_weight(r["net_weight"]),
                _fmt_rate(r["rate"]),
                credit_cell,
                debit_cell,
            ]
        )

    usable_width = A4[0] - 72
    col_widths = [
        usable_width * 0.15,
        usable_width * 0.15,
        usable_width * 0.18,
        usable_width * 0.12,
        usable_width * 0.20,
        usable_width * 0.20,
    ]

    ledger_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ececec")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfcfcf")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfcfcf")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
    ]
    for i, r in enumerate(rows, start=1):
        if r["kind"] == "Purchase":
            style_cmds.append(("TEXTCOLOR", (1, i), (1, i), green))
            style_cmds.append(("FONTNAME", (1, i), (1, i), "Helvetica-Bold"))
            style_cmds.append(("TEXTCOLOR", (4, i), (4, i), green))
            style_cmds.append(("FONTNAME", (4, i), (4, i), "Helvetica-Bold"))
        else:
            style_cmds.append(("TEXTCOLOR", (1, i), (1, i), red))
            style_cmds.append(("FONTNAME", (1, i), (1, i), "Helvetica-Bold"))
            style_cmds.append(("TEXTCOLOR", (5, i), (5, i), red))
            style_cmds.append(("FONTNAME", (5, i), (5, i), "Helvetica-Bold"))

    ledger_table.setStyle(TableStyle(style_cmds))
    elements.append(ledger_table)
    elements.append(Spacer(1, 20))

    summary_data = [
        ["Total Credit", f"Rs.{_fmt_inr(total_credit)}"],
        ["Total Debit", f"Rs.{_fmt_inr(total_debit)}"],
        ["Difference (Credit - Debit)", f"Rs.{_fmt_inr(difference)}"],
        ["Paid Amount", paid_label],
        ["Closing Balance", closing_label],
    ]
    summary_table = Table(summary_data, colWidths=[usable_width * 0.35, usable_width * 0.25])
    summary_table.hAlign = "RIGHT"
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, 3), colors.HexColor("#efefef")),
                ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#f8f8f8")),
                ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#fff8e8")),
                ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor("#d9edf7")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 3), 10),
                ("FONTSIZE", (0, 4), (-1, 4), 12),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfcfcf")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfcfcf")),
            ]
        )
    )
    elements.append(summary_table)

    doc.build(elements)
    buffer.seek(0)

    safe_name = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in (party.name or "party"))[:40]
    filename = f"party_ledger_{safe_name}_{from_date.strftime('%Y%m%d')}_{to_date.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/parties-balance-sheet")
async def generate_parties_balance_sheet(
    language: Optional[str] = Query("en", description="Party name language: en or ta"),
    db: AsyncSession = Depends(get_db),
):
    """All-parties Balance Sheet PDF: Party Name | Credit (To Pay) | Debit (To Receive)."""
    from datetime import datetime as dt

    lang = _normalize_lang(language)
    result = await db.execute(
        select(Party).where(Party.is_active == True).order_by(Party.name.asc())  # noqa: E712
    )
    parties = result.scalars().all()

    generated_on = dt.now().strftime("%d-%m-%Y")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "BalanceSheetTitle",
        parent=styles["Heading1"],
        alignment=1,
        fontSize=22,
        spaceAfter=6,
    )
    date_style = ParagraphStyle(
        "BalanceSheetDate",
        parent=styles["Normal"],
        alignment=1,
        fontSize=11,
        textColor=colors.HexColor("#555555"),
        spaceAfter=18,
    )

    elements = []
    elements.append(Paragraph("Balance Sheet", title_style))
    elements.append(Paragraph(f"Generate Date: {generated_on}", date_style))

    table_data = [["Party Name", "Credit (To Pay)", "Debit (To Receive)"]]
    total_credit = 0.0
    total_debit = 0.0

    for party in parties:
        bal = float(party.current_balance or 0)
        if bal > 0:
            credit = bal
            debit = 0.0
            credit_cell = f"Rs.{_fmt_inr(credit)}"
            debit_cell = "-"
        elif bal < 0:
            credit = 0.0
            debit = abs(bal)
            credit_cell = "-"
            debit_cell = f"Rs.{_fmt_inr(debit)}"
        else:
            credit = 0.0
            debit = 0.0
            credit_cell = "-"
            debit_cell = "-"

        total_credit += credit
        total_debit += debit
        table_data.append([_party_name_cell(_party_display_name(party, lang), lang, 10), credit_cell, debit_cell])

    table_data.append(
        [
            "Total",
            f"Rs.{_fmt_inr(total_credit)}",
            f"Rs.{_fmt_inr(total_debit)}",
        ]
    )

    usable_width = A4[0] - 72
    col_widths = [usable_width * 0.46, usable_width * 0.27, usable_width * 0.27]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    last_row = len(table_data) - 1
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ececec")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("FONTNAME", (0, 1), (-1, last_row - 1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("FONTNAME", (0, last_row), (-1, last_row), "Helvetica-Bold"),
                ("BACKGROUND", (0, last_row), (-1, last_row), colors.HexColor("#f0f4f2")),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfcfcf")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cfcfcf")),
                ("ROWBACKGROUNDS", (0, 1), (-1, last_row - 1), [colors.white, colors.HexColor("#fafafa")]),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"parties_balance_sheet_{dt.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
